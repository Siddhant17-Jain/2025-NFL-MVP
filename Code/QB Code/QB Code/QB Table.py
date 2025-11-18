import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook

# Directory containing the CSV files
data_dir = '/Users/siddhantjain/PycharmProjects/NFL MVP/QB Files/Separate CSV'

# List of teams we are interested in
teams = ['Joe Burrow','Josh Allen','Lamar Jackson','Jared Goff','Sam Darnold']

# List of the statistics files (this should match the files in your directory)
metrics = [
    "win_pct","pass_atts","td","passing_yards","comp_pct",
    "passer_rating","avg_epa","success_rate","total_wpa","qbr_total","int"
]

# Dictionary to map metric abbreviations to full names
metric_full_names = {
    "win_pct": "Win Percentage",
    "pass_atts": "Pass Attempts",
    "td": "Touchdowns",
    "passing_yards": "Passing Yards",
    "comp_pct": "Completion Percentage",
    "passer_rating": "Passer Rating",
    "avg_epa": "Average EPA",
    "success_rate": "Success Rate",
    "total_wpa": "Total Win Probability Added",
    "qbr_total": "QBR Total",
    "int": "Interceptions"
}


# Create an Excel workbook to store the table
wb = Workbook()
ws = wb.active
ws.title = "Team Rankings"

# Define color fills for different ranks
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
amber_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
orange_fill = PatternFill(start_color="FF8000", end_color="FF8000", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Write the header row
ws.append(["Statistic"] + teams)

# Loop through each metric file
for metric in metrics:
    # Define the file path for the metric
    file_path = os.path.join(data_dir, f"{metric.replace(' ', '_')}_Ranked.csv")

    # Load the metric CSV file
    df = pd.read_csv(file_path)

    # Get the full name for the metric
    full_metric_name = metric_full_names.get(metric, metric)  # Use the full name if available

    # Get the rankings for each team
    row_data = [full_metric_name]  # Start the row with the full metric name

    for team in teams:
        # Find the row for the team and get its value and rank in the current metric
        team_row = df[df["Player"] == team]

        if not team_row.empty:
            value = team_row[metric].values[0]  # Get the value for the statistic
            rank = team_row[f"{metric}_Rank"].values[0]  # Get the rank for the statistic

            # Determine the color based on rank
            if 1 <= rank <= 10:
                fill = green_fill
            elif 11 <= rank <= 20:
                fill = amber_fill
            elif 21 <= rank <= 26:
                fill = orange_fill
            else:
                fill = red_fill

            # Append the value and style the cell
            row_data.append((value, fill))
        else:
            # If the team is not found (error handling), add a placeholder value
            row_data.append(("N/A", None))

    # Write the row to the worksheet
    ws.append([value[0] for value in row_data])  # Write only the values to the cells

    # Apply the color fill to the cells
    for col_num, value in enumerate(row_data[1:], start=2):  # Start from the second column
        if value[1]:  # If the cell has a fill color
            ws.cell(row=ws.max_row, column=col_num).fill = value[1]

# Save the workbook to a file
output_file = '/Users/siddhantjain/Desktop/Team_Stats_Table.xlsx'
wb.save(output_file)

print("Excel file with full metric names and color-coded rankings created and saved to your desktop!")
